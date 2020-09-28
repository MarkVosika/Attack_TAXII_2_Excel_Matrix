# Import needed libraries
from stix2 import TAXIICollectionSource, Filter
from taxii2client.v20 import Server, Collection
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

#__________________________________________________________________________________________________________________________
# Enterprise attack source
collection = Collection("https://cti-taxii.mitre.org/stix/collections/95ecc380-afe9-11e4-9b6c-751b66dd541e/")
# Supply the TAXII2 collection to TAXIICollection
tc_source = TAXIICollectionSource(collection)

# Filter to only techniques
filt = Filter('type', '=', 'attack-pattern')
techniques = tc_source.query([filt])

#__________________________________________________________________________________________________________________________
# Generate list of all kill chain phases and put them in proper order
kc_list = []

for technique in techniques:
	for k,v in technique.items():
		if k == "kill_chain_phases":
			for i in v:
				for k,v in i.items():
					kc_list.append(v)

# Remove duplicates and sort
deduped_kc = sorted(list(dict.fromkeys(kc_list)))
deduped_kc.remove('mitre-attack')
reorder = [8,5,10,11,3,2,4,9,0,1]
deduped_kc = [deduped_kc[i] for i in reorder]

#__________________________________________________________________________________________________________________________
# Seperate platforms
windows = []
linux = []
mac = []
all_techniques = []

for technique in techniques:
    for k,v in technique.items():
    	if k == 'x_mitre_platforms':
    		all_techniques.append(technique)
    		if 'Windows' in v:
    			windows.append(technique)			
    		if "Linux" in v:
    			linux.append(technique)
    		if "macOS" in v:
    			mac.append(technique)

#__________________________________________________________________________________________________________________________
# Create worksheet tabs and variable to hold mitre data (list_of_lists)
count = 1

wb = Workbook()
sheet1 = wb['Sheet']
sheet1.title = 'All_Techniques'
sheet2 = wb.create_sheet('Windows')
sheet3 = wb.create_sheet('Linux')
sheet4 = wb.create_sheet('macOS')

list_of_lists = []

#__________________________________________________________________________________________________________________________
# Fuction to create a flatlist of a nested list, for each list in list_of_lists
def flattenNestedList(nestedList):
    ''' Converts a nested list to a flat list '''
    flatList = []
    # Iterate over all the elements in given list
    for elem in nestedList:
        # Check if type of element is list
        if isinstance(elem, list):
            # Extend the flat list by adding contents of this element (list)
            flatList.extend(flattenNestedList(elem))
        else:
            # Append the elemengt to the list
            flatList.append(elem)    
    return flatList

#__________________________________________________________________________________________________________________________
# Function to search each flatlist and return a count of occurences for input
def search_flatlist(list_to_search, technique_id):
	count = 0
	for i in list_to_search:
		if re.search(technique_id, str(i)):
			count += 1
	return count

#__________________________________________________________________________________________________________________________
# Function to parse the mitre data and extract the Technique ID and Name
# It then calls the other functions to remove any parent techniques (using technique ID) if a sub technique is present
# Lastly it exports everything to excel, recreating the Mitre Attack framework
def parse_json(platform):
	global count
	#create empty list for each kill chain phase:

	list_of_lists = []

	for i in deduped_kc:
		list_name = []
		list_of_lists.append(list_name)

	for lst in platform:
		for technique in lst:
			for i in technique['external_references']:
				for k,v in i.items():
					if k == 'external_id' and v.startswith('T'): #extracts the technique ID
						t_id = v
			for index in enumerate(deduped_kc):
				for k,v in technique.items():
					if k == "kill_chain_phases":
						for i in v:
							for k,v in i.items():
								if index[1] == v:
									#if len(t_id) < 6: #uncomment this line to remove subtechniques
										list_of_lists[index[0]].append([t_id, technique["name"]])

	for lst in list_of_lists:
		flattenedlist = []
		flattenedlist = flattenNestedList(lst)
		for i in lst:
			if search_flatlist(flattenedlist,i[0]) > 1:
				lst.remove(i)

	# Write to excel file all the data, making first row bold
	bold_font = Font(bold = True)		#set bold font variable
	wrap_text = Alignment(wrap_text=True)

	if count == 1:
		current_sheet = sheet1
	elif count == 2:
		current_sheet = sheet2
	elif count == 3:
		current_sheet = sheet3
	elif count == 4:
		current_sheet = sheet4

	current_sheet.append(deduped_kc)
	for cell in current_sheet["1:1"]:
		cell.font = bold_font
		cell.alignment = wrap_text
	
	# This part is how the Excel columns are built, incrementing character and count
	column = '@'

	for lst in list_of_lists:
		column = chr(ord(column) + 1)
		row = 2
		for technique in lst:
			current_sheet[column + str(row)] = technique[1]
			#current_sheet[column + str(row)].alignment = wrap_text
			row += 1 

	count += 1

	wb.save('file_name.xlsx')							

#__________________________________________________________________________________________________________________________
# Call main function with each platform type
parse_json([all_techniques])
parse_json([windows]) 
parse_json([linux])
parse_json([mac])
